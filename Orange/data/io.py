import contextlib
import csv
import locale
import pickle
import re
import sys
import warnings
from typing import List, Iterable

from functools import lru_cache
from importlib import import_module
from itertools import chain

from os import path, remove
from tempfile import NamedTemporaryFile
from urllib.parse import urlparse, urlsplit, urlunsplit, \
    unquote as urlunquote, quote
from urllib.request import urlopen, Request
from pathlib import Path

import numpy as np

import xlrd
import xlsxwriter
import openpyxl

from Orange.data import _io, Table, Domain, ContinuousVariable
from Orange.data import Compression, open_compressed, detect_encoding, \
    isnastr, guess_data_type, sanitize_variable
from Orange.data.io_base import FileFormatBase, Flags, DataTableMixin, PICKLE_PROTOCOL

from Orange.util import flatten


# Support values longer than 128K (i.e. text contents features)
csv.field_size_limit(100*1024*1024)

__all__ = ["Flags", "FileFormat"]


Compression = Compression
open_compressed = open_compressed
detect_encoding = detect_encoding
isnastr = isnastr
guess_data_type = guess_data_type
sanitize_variable = sanitize_variable
Flags = Flags
FileFormatMeta = type(FileFormatBase)


class FileFormat(FileFormatBase):
    """
    Subclasses set the following attributes and override the following methods:

        EXTENSIONS = ('.ext1', '.ext2', ...)
        DESCRIPTION = 'human-readable file format description'
        SUPPORT_COMPRESSED = False
        SUPPORT_SPARSE_DATA = False

        def read(self):
            ...  # load headers, data, ...
            return self.data_table(data, headers)

        @classmethod
        def write_file(cls, filename, data):
            ...
            self.write_headers(writer.write, data)
            writer.writerows(data)

    Wrapper FileFormat.data_table() returns Orange.data.Table from `data`
    iterable (list (rows) of lists of values (cols)).
    """

    # Priority when multiple formats support the same extension. Also
    # the sort order in file open/save combo boxes. Lower is better.
    PRIORITY = 10000
    OPTIONAL_TYPE_ANNOTATIONS = False
    SUPPORT_COMPRESSED = False
    SUPPORT_SPARSE_DATA = False

    def __init__(self, filename):
        """
        Parameters
        ----------
        filename : str
            name of the file to open
        """
        self.filename = filename
        self.sheet = None

    @property
    def sheets(self) -> List:
        """FileFormats with a notion of sheets should override this property
        to return a list of sheet names in the file.

        Returns
        -------
        a list of sheet names
        """
        return []

    def select_sheet(self, sheet):
        """Select sheet to be read

        Parameters
        ----------
        sheet : str
            sheet name
        """
        self.sheet = sheet


def class_from_qualified_name(format_name):
    """ File format class from qualified name. """
    elements = format_name.split(".")
    m = import_module(".".join(elements[:-1]))
    return getattr(m, elements[-1])


class CSVReader(FileFormat, DataTableMixin):
    """Reader for comma separated files"""

    EXTENSIONS = ('.csv',)
    DESCRIPTION = 'Comma-separated values'
    DELIMITERS = ',;:\t$ '
    SUPPORT_COMPRESSED = True
    SUPPORT_SPARSE_DATA = False
    PRIORITY = 20
    OPTIONAL_TYPE_ANNOTATIONS = True

    def read(self):
        for encoding in (lambda: ('us-ascii', None),                 # fast
                         lambda: (detect_encoding(self.filename), None),  # precise
                         lambda: (locale.getpreferredencoding(False), None),
                         lambda: (sys.getdefaultencoding(), None),   # desperate
                         lambda: ('utf-8', None),                    # ...
                         lambda: ('utf-8', 'ignore')):               # fallback
            encoding, errors = encoding()
            # Clear the error flag for all except the last check, because
            # the error of second-to-last check is stored and shown as warning in owfile
            if errors != 'ignore':
                error = ''
            with self.open(self.filename, mode='rt', newline='',
                           encoding=encoding, errors=errors) as file:
                # Sniff the CSV dialect (delimiter, quotes, ...)
                try:
                    dialect = csv.Sniffer().sniff(
                        # Take first couple of *complete* lines as sample
                        ''.join(file.readline() for _ in range(10)),
                        self.DELIMITERS)
                    delimiter = dialect.delimiter
                    quotechar = dialect.quotechar
                except UnicodeDecodeError as e:
                    error = e
                    continue
                except csv.Error:
                    delimiter = self.DELIMITERS[0]
                    quotechar = csv.excel.quotechar

                file.seek(0)
                try:
                    reader = csv.reader(
                        file, delimiter=delimiter, quotechar=quotechar,
                        skipinitialspace=True,
                    )
                    data = self.data_table(reader)

                    # TODO: Name can be set unconditionally when/if
                    # self.filename will always be a string with the file name.
                    # Currently, some tests pass StringIO instead of
                    # the file name to a reader.
                    if isinstance(self.filename, str):
                        data.name = path.splitext(
                            path.split(self.filename)[-1])[0]
                    if error and isinstance(error, UnicodeDecodeError):
                        pos, endpos = error.args[2], error.args[3]
                        warning = ('Skipped invalid byte(s) in position '
                                   '{}{}').format(pos,
                                                  ('-' + str(endpos)) if (endpos - pos) > 1 else '')
                        warnings.warn(warning)
                    self.set_table_metadata(self.filename, data)
                    return data
                except Exception as e:
                    error = e
                    continue
        raise ValueError('Cannot parse dataset {}: {}'.format(self.filename, error)) from error

    @classmethod
    def write_file(cls, filename, data, with_annotations=True):
        with cls.open(filename, mode='wt', newline='', encoding='utf-8') as file:
            writer = csv.writer(file, delimiter=cls.DELIMITERS[0])
            cls.write_headers(writer.writerow, data, with_annotations)
            cls.write_data(writer.writerow, data)
            cls.write_table_metadata(filename, data)


class TabReader(CSVReader):
    """Reader for tab separated files"""
    EXTENSIONS = ('.tab', '.tsv')
    DESCRIPTION = 'Tab-separated values'
    DELIMITERS = '\t'
    PRIORITY = 10


class PickleReader(FileFormat):
    """Reader for pickled Table objects"""
    EXTENSIONS = ('.pkl', '.pickle')
    DESCRIPTION = 'Pickled Orange data'
    SUPPORT_COMPRESSED = True
    SUPPORT_SPARSE_DATA = True

    def read(self):
        with self.open(self.filename, 'rb') as f:
            table = pickle.load(f)
            if not isinstance(table, Table):
                raise TypeError("file does not contain a data table")
            else:
                return table

    @classmethod
    def write_file(cls, filename, data):
        with cls.open(filename, 'wb') as f:
            pickle.dump(data, f, protocol=PICKLE_PROTOCOL)


class BasketReader(FileFormat):
    """Reader for basket (sparse) files"""
    EXTENSIONS = ('.basket', '.bsk')
    DESCRIPTION = 'Basket file'
    SUPPORT_SPARSE_DATA = True

    def read(self):
        def constr_vars(inds):
            if inds:
                return [ContinuousVariable(x.decode("utf-8")) for _, x in
                        sorted((ind, name) for name, ind in inds.items())]

        X, Y, metas, attr_indices, class_indices, meta_indices = \
            _io.sparse_read_float(self.filename.encode(sys.getdefaultencoding()))

        attrs = constr_vars(attr_indices)
        classes = constr_vars(class_indices)
        meta_attrs = constr_vars(meta_indices)
        domain = Domain(attrs, classes, meta_attrs)
        table = Table.from_numpy(
            domain, attrs and X, classes and Y, metas and meta_attrs)
        table.name = path.splitext(path.split(self.filename)[-1])[0]
        return table


class _BaseExcelReader(FileFormat, DataTableMixin):
    """Base class for reading excel files"""
    SUPPORT_COMPRESSED = False
    SUPPORT_SPARSE_DATA = False

    def __init__(self, filename):
        super().__init__(filename=filename)
        self._workbook = None

    def get_cells(self) -> Iterable:
        raise NotImplementedError

    def read(self):
        try:
            cells = self.get_cells()
            table = self.data_table(cells)
            table.name = path.splitext(path.split(self.filename)[-1])[0]
            if self.sheet and len(self.sheets) > 1:
                table.name = '-'.join((table.name, self.sheet))
        except Exception:
            raise IOError("Couldn't load spreadsheet from " + self.filename)
        return table


class ExcelReader(_BaseExcelReader):
    """Reader for .xlsx files"""
    EXTENSIONS = ('.xlsx',)
    DESCRIPTION = 'Microsoft Excel spreadsheet'
    ERRORS = ("#VALUE!", "#DIV/0!", "#REF!", "#NUM!", "#NULL!", "#NAME?")
    OPTIONAL_TYPE_ANNOTATIONS = True

    def __init__(self, filename):
        super().__init__(filename)
        self.sheet = self.workbook.active.title

    @property
    def workbook(self) -> openpyxl.Workbook:
        if not self._workbook:
            with warnings.catch_warnings():
                # We don't care about extensions, but we hate warnings
                warnings.filterwarnings(
                    "ignore",
                    ".*extension is not supported and will be removed.*",
                    UserWarning)
                self._workbook = openpyxl.load_workbook(self.filename,
                                                    data_only=True)
        return self._workbook

    @property
    @lru_cache(1)
    def sheets(self) -> List:
        return self.workbook.sheetnames if self.workbook else []

    def get_cells(self) -> Iterable:
        def str_(x):
            return str(x) if x is not None and x not in ExcelReader.ERRORS \
                else ""

        sheet = self._get_active_sheet()
        min_col = sheet.min_column
        max_col = sheet.max_column
        cells = ([str_(cell.value) for cell in row[min_col - 1: max_col]]
                 for row in sheet.iter_rows(sheet.min_row, sheet.max_row + 1))
        return filter(any, cells)

    def _get_active_sheet(self) -> openpyxl.worksheet.worksheet.Worksheet:
        if self.sheet:
            return self.workbook[self.sheet]
        else:
            return self.workbook.active

    @classmethod
    def write_file(cls, filename, data, with_annotations=False):
        vars = list(chain((ContinuousVariable('_w'),) if data.has_weights() else (),
                          data.domain.attributes,
                          data.domain.class_vars,
                          data.domain.metas))
        formatters = [cls.formatter(v) for v in vars]
        zipped_list_data = zip(data.W if data.W.ndim > 1 else data.W[:, np.newaxis],
                               data.X,
                               data.Y if data.Y.ndim > 1 else data.Y[:, np.newaxis],
                               data.metas)
        names = cls.header_names(data)
        headers = (names,)
        if with_annotations:
            types = cls.header_types(data)
            flags = cls.header_flags(data)
            headers = (names, types, flags)

        workbook = xlsxwriter.Workbook(filename)
        sheet = workbook.add_worksheet()

        for r, parts in enumerate(headers):
            for c, part in enumerate(parts):
                sheet.write(r, c, part)
        for i, row in enumerate(zipped_list_data, len(headers)):
            for j, (fmt, v) in enumerate(zip(formatters, flatten(row))):
                sheet.write(i, j, fmt(v))
        workbook.close()


class XlsReader(_BaseExcelReader):
    """Reader for .xls files"""
    EXTENSIONS = ('.xls',)
    DESCRIPTION = 'Microsoft Excel 97-2004 spreadsheet'

    def __init__(self, filename):
        super().__init__(filename)
        self.sheet = self.workbook.sheet_by_index(0).name

    @property
    def workbook(self) -> xlrd.Book:
        if not self._workbook:
            self._workbook = xlrd.open_workbook(self.filename, on_demand=True)
        return self._workbook

    @property
    @lru_cache(1)
    def sheets(self) -> List:
        return self.workbook.sheet_names() if self.workbook else []

    def get_cells(self) -> Iterable:
        def str_(cell):
            return "" if cell.ctype == xlrd.XL_CELL_ERROR else str(cell.value)

        sheet = self._get_active_sheet()
        first_row = next(i for i in range(sheet.nrows)
                         if any(sheet.row_values(i)))
        first_col = next(i for i in range(sheet.ncols)
                         if sheet.cell_value(first_row, i))
        row_len = sheet.row_len(first_row)
        return filter(any, ([str_(sheet.cell(row, col))
                             if col < sheet.row_len(row) else ''
                             for col in range(first_col, row_len)]
                            for row in range(first_row, sheet.nrows)))

    def _get_active_sheet(self) -> xlrd.sheet.Sheet:
        if self.sheet:
            return self.workbook.sheet_by_name(self.sheet)
        else:
            return self.workbook.sheet_by_index(0)


class DotReader(FileFormat):
    """Writer for dot (graph) files"""
    EXTENSIONS = ('.dot', '.gv')
    DESCRIPTION = 'Dot graph description'
    SUPPORT_COMPRESSED = True
    SUPPORT_SPARSE_DATA = False

    @classmethod
    def write_graph(cls, filename, graph):
        from sklearn import tree
        tree.export_graphviz(graph, out_file=cls.open(filename, 'wt'))

    @classmethod
    def write(cls, filename, tree):
        if type(tree) == dict:
            tree = tree['tree']
        cls.write_graph(filename, tree)


class UrlReader(FileFormat):
    def __init__(self, filename):
        filename = filename.strip()
        if not urlparse(filename).scheme:
            filename = 'http://' + filename

        # Fully support URL with query or fragment like http://filename.txt?a=1&b=2#c=3
        def quote_byte(b):
            return chr(b) if b < 0x80 else '%{:02X}'.format(b)

        filename = ''.join(map(quote_byte, filename.encode("utf-8")))

        super().__init__(filename)

    @staticmethod
    def urlopen(url):
        req = Request(
            url,
            # Avoid 403 error with servers that dislike scrapers
            headers={'User-Agent': 'Mozilla/5.0 (X11; Linux) Gecko/20100101 Firefox/'})
        return urlopen(req, timeout=10)

    def read(self):
        self.filename = self._trim(self._resolve_redirects(self.filename))
        with contextlib.closing(self.urlopen(self.filename)) as response:
            name = self._suggest_filename(response.headers['content-disposition'])
            # using Path since splitext does not extract more extensions
            extension = ''.join(Path(name).suffixes)  # get only file extension
            with NamedTemporaryFile(suffix=extension, delete=False) as f:
                f.write(response.read())
                # delete=False is a workaround for https://bugs.python.org/issue14243

            reader = self.get_reader(f.name)
            data = reader.read()
            remove(f.name)
        # Override name set in from_file() to avoid holding the temp prefix
        data.name = path.splitext(name)[0]
        data.origin = self.filename
        return data

    def _resolve_redirects(self, url):
        # Resolve (potential) redirects to a final URL
        with contextlib.closing(self.urlopen(url)) as response:
            return response.url

    @classmethod
    def _trim(cls, url):
        URL_TRIMMERS = (
            cls._trim_googlesheet,
            cls._trim_dropbox,
        )
        for trim in URL_TRIMMERS:
            try:
                url = trim(url)
            except ValueError:
                continue
            else:
                break
        return url

    @staticmethod
    def _trim_googlesheet(url):
        match = re.match(r'(?:https?://)?(?:www\.)?'
                         r'docs\.google\.com/spreadsheets/d/'
                         r'(?P<workbook_id>[-\w_]+)'
                         r'(?:/.*?gid=(?P<sheet_id>\d+).*|.*)?',
                         url, re.IGNORECASE)
        try:
            workbook, sheet = match.group('workbook_id'), match.group('sheet_id')
            if not workbook:
                raise ValueError
        except (AttributeError, ValueError):
            raise ValueError
        url = 'https://docs.google.com/spreadsheets/d/{}/export?format=tsv'.format(workbook)
        if sheet:
            url += '&gid=' + sheet
        return url

    @staticmethod
    def _trim_dropbox(url):
        parts = urlsplit(url)
        if not parts.netloc.endswith('dropbox.com'):
            raise ValueError
        return urlunsplit(parts._replace(query='dl=1'))

    def _suggest_filename(self, content_disposition):
        default_name = re.sub(r'[\\:/]', '_', urlparse(self.filename).path)

        # See https://tools.ietf.org/html/rfc6266#section-4.1
        matches = re.findall(r"filename\*?=(?:\"|.{0,10}?'[^']*')([^\"]+)",
                             content_disposition or '')
        return urlunquote(matches[-1]) if matches else default_name
