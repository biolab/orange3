import numpy as np
import openpyxl


def read_matrix(filename, sheet_name=None):
    sheet = _get_sheet(filename, sheet_name)
    cells, empty_cols, empty_rows = _non_empty_cells(sheet)
    if cells.dtype in (float, np.float64, np.float32):
        return cells, None, None, 1

    col_labels = _get_labels(cells[0])
    row_labels = _get_labels(cells[:, 0])
    if col_labels and row_labels:
        col_labels = col_labels[1:]
        row_labels = row_labels[1:]
    cells = cells[int(bool(col_labels)):, int(bool(row_labels)):]
    matrix = _matrix_from_cells(
        cells, empty_cols + bool(row_labels), empty_rows + bool(col_labels))
    return matrix, row_labels, col_labels, 1


def _get_sheet(filename, sheet_name):
    workbook = openpyxl.load_workbook(filename, data_only=True)
    if sheet_name is None:
        return workbook.active
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"No such sheet: {sheet_name}")
    return workbook.worksheets[workbook.sheetnames.index(sheet_name)]


def _non_empty_cells(sheet):
    """
    Reported sheet.max_column and sheet.max_row may be too large, so
    we must read all cells from (supposedly) used region and trim it.
    Since we must remove empty rows and columns at the end anywey, we
    are kind to users and also remove leading empty rows and columns.

    Returns:
         - np.array with non-empty part of the sheet
         - number of empty columns to the left
         - number of empty rows above
    """
    def raise_empty():
        raise ValueError("empty sheet")

    cells = np.array([[cell.value for cell in row] for row in sheet.rows])
    # Quick route out for any large table of numbers
    if not cells.size:
        raise_empty()
    if np.can_cast(cells.dtype, float):
        return cells.astype(float), 0, 0

    # compares by array cells, pylint: disable=singleton-comparison
    nonempty = cells != None
    offsets = []
    for _ in range(2):
        nonem = np.cumsum(np.any(nonempty, axis=1))
        if nonem[-1] == 0:
            raise_empty()
        mask = (nonem > 0) & (nonem < nonem[-1])
        # The last element that increased cumsum is also non-empty
        mask[1:] |= mask[:-1]
        offsets.append(np.sum(nonem == 0))
        cells = cells[mask]
        cells = cells.T
        nonempty = nonempty.T
    return cells, *offsets


def _get_labels(labels):
    try:
        for label in labels[1:]:
            # pylint: disable=expression-not-assigned
            label is None or float(label)
    except ValueError:
        return ["?" if label is None else str(label)
                for label in labels]
    else:
        return None


def _matrix_from_cells(cells, row_offset, col_offset):
    matrix = np.full(cells.shape, np.nan)
    for y, row in enumerate(cells):
        for x, value in enumerate(row):
            if value is None:
                continue
            if isinstance(value, (int, float)):
                matrix[y, x] = value
                continue
            try:
                # Triggers AttributeError, if value is not a string
                if not value.strip():
                    continue
                # Triggers ValueError if not a number
                matrix[y, x] = float(value)
            except (AttributeError, ValueError):
                raise ValueError(
                    "invalid data in cell "
                    f"{openpyxl.utils.get_column_letter(x + col_offset + 1)}"
                    f"{y + row_offset + 1}") from None
    return matrix


def write_matrix(matrix: "DistMatrix", filename):
    wb = openpyxl.Workbook()
    sheet = wb.active
    row_labels = matrix.get_labels(matrix.row_items)
    col_labels = (matrix.col_items is not matrix.row_items) \
                 and matrix.get_labels(matrix.col_items)
    has_row_labels = row_labels is not None
    has_col_labels = col_labels is not None and col_labels is not False
    row_off = 1 + int(has_col_labels)
    col_off = 1 + int(has_row_labels)

    if has_row_labels:
        for row, label in enumerate(row_labels, start=row_off):
            sheet.cell(row, 1).value = label
    if has_col_labels:
        for col, label in enumerate(col_labels, start=col_off):
            sheet.cell(1, col).value = label
    symmetric = matrix.is_symmetric()
    has_diagonal = int(np.any(np.diag(matrix) != 0))
    for y in range(matrix.shape[0]):
        for x in range(y + has_diagonal if symmetric else matrix.shape[1]):
            sheet.cell(y + row_off, x + col_off).value = matrix[y, x]

    wb.save(filename)
