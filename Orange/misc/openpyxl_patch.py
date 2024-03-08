import warnings

from openpyxl.cell import MergedCell
from openpyxl.comments.comment_sheet import CommentSheet
from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing
from openpyxl.packaging.relationship import (
    get_rels_path,
    RelationshipList,
    get_dependents,
)
from openpyxl.pivot.table import TableDefinition
from openpyxl.reader.drawings import find_images
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet._reader import WorksheetReader
from openpyxl.xml.constants import COMMENTS_NS
from openpyxl.xml.functions import fromstring
from openpyxl.worksheet.table import Table


# temporary fix for file not closed issue until openpyxl prepare release
# https://foss.heptapod.net/openpyxl/openpyxl/-/merge_requests/
# 436#7922bd5f66e11e4ca4539f093b2680a25c1f80db
def read_worksheets(self):
    # pylint: disable=too-many-branches
    comment_warning = (
        "Cell '{0}':{1} is part of a merged range but has a comment which will "
        "be removed because merged cells cannot contain any data."
    )
    for sheet, rel in self.parser.find_sheets():
        if rel.target not in self.valid_files:
            continue

        if "chartsheet" in rel.Type:
            self.read_chartsheet(sheet, rel)
            continue

        rels_path = get_rels_path(rel.target)
        rels = RelationshipList()
        if rels_path in self.valid_files:
            rels = get_dependents(self.archive, rels_path)

        if self.read_only:
            ws = ReadOnlyWorksheet(self.wb, sheet.name, rel.target, self.shared_strings)
            ws.sheet_state = sheet.state
            self.wb._sheets.append(ws)  # pylint: disable=protected-access
            continue
        fh = self.archive.open(rel.target)
        ws = self.wb.create_sheet(sheet.name)
        ws._rels = rels  # pylint: disable=protected-access
        ws_parser = WorksheetReader(
            ws, fh, self.shared_strings, self.data_only, self.rich_text
        )
        ws_parser.bind_all()
        fh.close()

        # assign any comments to cells
        for r in rels.find(COMMENTS_NS):
            src = self.archive.read(r.target)
            comment_sheet = CommentSheet.from_tree(fromstring(src))
            for ref, comment in comment_sheet.comments:
                try:
                    ws[ref].comment = comment
                except AttributeError:
                    c = ws[ref]
                    if isinstance(c, MergedCell):
                        warnings.warn(comment_warning.format(ws.title, c.coordinate))
                        continue

        # preserve link to VML file if VBA
        if self.wb.vba_archive and ws.legacy_drawing:
            ws.legacy_drawing = rels.get(ws.legacy_drawing).target
        else:
            ws.legacy_drawing = None

        for t in ws_parser.tables:
            src = self.archive.read(t)
            xml = fromstring(src)
            table = Table.from_tree(xml)
            ws.add_table(table)

        #pylint: disable=protected-access
        drawings = rels.find(SpreadsheetDrawing._rel_type)
        for rel in drawings:
            charts, images = find_images(self.archive, rel.target)
            for c in charts:
                ws.add_chart(c, c.anchor)
            for im in images:
                ws.add_image(im, im.anchor)

        pivot_rel = rels.find(TableDefinition.rel_type)
        pivot_caches = self.parser.pivot_caches
        for r in pivot_rel:
            pivot_path = r.Target
            src = self.archive.read(pivot_path)
            tree = fromstring(src)
            pivot = TableDefinition.from_tree(tree)
            pivot.cache = pivot_caches[pivot.cacheId]
            ws.add_pivot(pivot)

        ws.sheet_state = sheet.state
